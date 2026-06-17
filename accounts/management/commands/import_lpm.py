"""Import des donnees reelles LPM depuis le classeur Excel (LPM_modele_donnees.xlsx).

Usage :
    python manage.py import_lpm                 # import (upsert) depuis le fichier par defaut
    python manage.py import_lpm --flush         # VIDE toute la base puis importe (remplace la demo)
    python manage.py import_lpm --dry-run       # simulation : lit et verifie, n'ecrit rien
    python manage.py import_lpm --file chemin.xlsx --password Lpm@2026

Onglets attendus : Departements, Employes, Clients, Projets, Magasin
(voir l'onglet "Lisez-moi" du modele pour les colonnes et valeurs autorisees).
"""

import os
import unicodedata
from datetime import date, datetime

from django.conf import settings
from django.core.management import call_command
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

# --- Tables de correspondance libelle FR -> code interne ---
ROLE_MAP = {
    "directeur general": "CEO", "ceo": "CEO", "dg": "CEO",
    "responsable rh": "RH", "rh": "RH",
    "responsable de service": "MANAGER", "responsable": "MANAGER", "manager": "MANAGER",
    "employe": "EMPLOYE", "employe(e)": "EMPLOYE",
    "stagiaire": "STAGIAIRE",
}
CONTRACT_MAP = {"cdi": "CDI", "cdd": "CDD", "stage": "STAGE",
                "temporaire": "TEMP", "temp": "TEMP", "mission": "TEMP"}
GENDER_MAP = {"h": "M", "homme": "M", "m": "M", "f": "F", "femme": "F"}
PROJ_KIND_MAP = {"campagne": "CAMPAIGN", "evenement": "EVENT", "autre": "OTHER"}
PROJ_STATUS_MAP = {"termine": "DONE", "en cours": "ACTIVE", "a demarrer": "PLANNED",
                   "planifie": "PLANNED", "en pause": "ON_HOLD", "annule": "CANCELLED"}
CLIENT_KIND_MAP = {"client": "CLIENT", "prospect": "PROSPECT"}
STOCK_CAT_MAP = {
    "evenement": "EVENEMENT", "technique / son-lumiere": "TECHNIQUE", "technique": "TECHNIQUE",
    "informatique": "IT", "mobilier": "MOBILIER", "consommable": "CONSOMMABLE",
    "fournitures bureau": "BUREAUTIQUE", "equipement general": "EQUIPMENT", "autre": "OTHER",
}
STOCK_STATE_MAP = {"neuf": "NEW", "bon etat": "GOOD", "use": "USED",
                   "en reparation": "REPAIR", "hors service": "OUT_OF_SERVICE", "hs": "OUT_OF_SERVICE"}

# Lignes d'exemple du modele a ignorer si elles ont ete laissees telles quelles.
EXAMPLE_EMAILS = {"d.saksak@lpmconsulting.cm"}
EXAMPLE_PROJECTS = {"campagne nescafe 3en1"}
DG_NAME = "Direction Générale"


def _ascii(s):
    """Minuscule sans accents pour comparer les libelles de maniere robuste."""
    s = "" if s in (None, "") else str(s).strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _norm_name(s):
    """Clef de rapprochement d'une personne : 'Prenom NOM' -> 'prenom nom'."""
    return " ".join(_ascii(s).split())


def _to_date(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_int(v, default=0):
    if v in (None, ""):
        return default
    try:
        return int(float(str(v).replace(" ", "").replace(" ", "")))
    except (ValueError, TypeError):
        return default


def _split(v):
    """Decoupe une cellule multi-valeurs (';' ou ',')."""
    if not v:
        return []
    return [p.strip() for p in str(v).replace(",", ";").split(";") if p.strip()]


class Command(BaseCommand):
    help = "Importe les donnees reelles LPM depuis le classeur Excel."

    def add_arguments(self, parser):
        parser.add_argument("--file", default=None, help="Chemin du classeur (.xlsx).")
        parser.add_argument("--magasin", default=None,
                            help="Classeur magasin separe (colonnes MATERIELS/QUANTITE/ACTION).")
        parser.add_argument("--password", default="Lpm@2026", help="Mot de passe initial des comptes.")
        parser.add_argument("--flush", action="store_true",
                            help="Vide entierement la base avant l'import (remplace la demo).")
        parser.add_argument("--dry-run", action="store_true",
                            help="Simulation : lit et verifie, n'ecrit rien.")

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError("openpyxl est requis : .venv\\Scripts\\python.exe -m pip install openpyxl")

        path = opts["file"] or os.path.join(settings.BASE_DIR, "LPM_modele_donnees.xlsx")
        if not os.path.exists(path):
            raise CommandError(f"Fichier introuvable : {path}")
        self.pwd = opts["password"]
        self.dry = opts["dry_run"]
        wb = load_workbook(path, data_only=True)

        rows = {name: self._rows(wb, name) for name in
                ("Departements", "Employes", "Clients", "Projets", "Magasin")}

        # Magasin : fichier separe optionnel (colonnes MATERIELS/QUANTITE/ACTION).
        self.magasin_simple = []
        mag_path = opts["magasin"]
        if mag_path:
            if not os.path.exists(mag_path):
                raise CommandError(f"Fichier magasin introuvable : {mag_path}")
            mwb = load_workbook(mag_path, data_only=True)
            self.magasin_simple = self._rows(mwb, mwb.sheetnames[0])

        self.stdout.write(self.style.MIGRATE_HEADING(
            f"Lecture de {os.path.basename(path)} :"))
        for name, rs in rows.items():
            self.stdout.write(f"  - {name}: {len(rs)} ligne(s)")
        if mag_path:
            self.stdout.write(f"  - Magasin (fichier separe): {len(self.magasin_simple)} ligne(s)")

        if self.dry:
            self._dry_run(rows)
            return

        if opts["flush"]:
            self.stdout.write(self.style.WARNING("Flush de la base (remplacement complet)…"))
            call_command("flush", "--no-input")

        with transaction.atomic():
            self._ensure_reference()
            self._import(rows)
        self.stdout.write(self.style.SUCCESS("\n[OK] Import termine."))

    # ------------------------------------------------------------------ #
    def _rows(self, wb, sheet):
        """Liste de dicts {entete: valeur} pour les lignes non vides (apres l'entete)."""
        if sheet not in wb.sheetnames:
            return []
        ws = wb[sheet]
        headers = [(_ascii(c.value), c.value) for c in ws[1]]
        out = []
        for row in ws.iter_rows(min_row=2):
            values = [c.value for c in row]
            if not any(v not in (None, "") for v in values):
                continue
            d = {}
            for (akey, _orig), val in zip(headers, values):
                d[akey] = val.strip() if isinstance(val, str) else val
            out.append(d)
        return out

    @staticmethod
    def _s(v):
        """Valeur cellule -> texte propre (gere les nombres saisis sans apostrophe)."""
        if v in (None, ""):
            return ""
        if isinstance(v, float) and v.is_integer():
            v = int(v)
        return str(v).strip()

    def _gs(self, d, *aliases):
        """Raccourci : recupere une colonne et la renvoie en texte propre."""
        return self._s(self._get(d, *aliases))

    @staticmethod
    def _get(d, *aliases):
        """Recupere une valeur par nom de colonne, du plus precis au plus large.

        Priorite : egalite exacte > debut de chaine > contenu. Evite qu'un alias
        court ('nom') ne capte une autre colonne ('prenom')."""
        cleaned = {k: k.rstrip(" *").strip() for k in d}  # "nom *" -> "nom"
        for a in aliases:                       # 1. egalite exacte
            for k, ck in cleaned.items():
                if ck == a:
                    return d[k]
        for a in aliases:                       # 2. debut de chaine
            for k, ck in cleaned.items():
                if ck.startswith(a):
                    return d[k]
        for a in aliases:                       # 3. contenu
            for k, ck in cleaned.items():
                if a in ck:
                    return d[k]
        return None

    def _username_for(self, first, last, taken):
        base = (_ascii(first)[:1] + "." + _ascii(last).replace(" ", "")) or "user"
        base = base.strip(".") or "user"
        u, i = base, 1
        while u in taken:
            i += 1
            u = f"{base}{i}"
        taken.add(u)
        return u

    # ------------------------------------------------------------------ #
    def _ensure_reference(self):
        """Donnees de reference indispensables : admin + types de conge + lieu."""
        from accounts.models import Role, User
        admin, created = User.objects.get_or_create(
            username="admin",
            defaults=dict(first_name="Super", last_name="Admin",
                          email="admin@lpmconsulting.cm", role=Role.ADMIN,
                          is_staff=True, is_superuser=True,
                          secondary_roles="CEO,RH,MANAGER,EMPLOYE"))
        if created:
            admin.set_password(self.pwd)
            admin.save()
        from conges.models import LeaveType
        for name, code, paid, deduct, days, ref, color in [
            ("Conge annuel paye", "ANNUEL", True, True, 18, "Art. 89 Code du Travail", "#0073DE"),
            ("Conge maladie", "MALADIE", True, False, 0, "Certificat medical", "#f59e0b"),
            ("Conge de maternite", "MATERNITE", True, False, 98, "Art. 84 - 14 semaines", "#ec4899"),
            ("Conge de paternite", "PATERNITE", True, False, 3, "Convention collective", "#6366f1"),
            ("Permission exceptionnelle", "EXCEPT", True, False, 0, "Evenement familial", "#10b981"),
            ("Conge sans solde", "SANS_SOLDE", False, False, 0, "Accord employeur", "#64748b"),
        ]:
            LeaveType.objects.get_or_create(code=code, defaults=dict(
                name=name, is_paid=paid, deducts_balance=deduct,
                default_days=days, legal_reference=ref, color=color))

    def _resolve_person(self, name):
        """Retrouve un utilisateur par 'Prenom Nom', tolerant a l'ordre des tokens :
        essaie le nom complet, puis le nom de famille (depart./fiche)."""
        key = _norm_name(name)
        if not key:
            return None
        if key in self.user_by_name:
            return self.user_by_name[key]
        tokens = set(key.split())
        cands = [u for ln, us in self.user_by_last.items() if ln in tokens for u in us]
        if len(cands) == 1:
            return cands[0]
        for u in cands:  # plusieurs memes noms : on departage par le prenom
            if _ascii(u.first_name).split() and _ascii(u.first_name).split()[0] in tokens:
                return u
        return cands[0] if cands else None

    def _import(self, rows):
        from accounts.models import Role, User
        from employees.models import Department, Employee, Position

        # ---- 1. Departements + departement chapeau "Direction Generale" ----
        dept_by_name = {}
        for d in rows["Departements"]:
            name = self._gs(d, "nom")
            if not name:
                continue
            obj, _ = Department.objects.get_or_create(name=name, defaults={
                "code": self._gs(d, "code")[:10]})
            obj.code = (self._gs(d, "code") or obj.code or "")[:10]
            obj.save()
            dept_by_name[_norm_name(name)] = obj
        dg, _ = Department.objects.get_or_create(name=DG_NAME, defaults={"code": "DG"})
        dept_by_name[_norm_name(DG_NAME)] = dg
        # Rattachement : chaque departement -> son parent ("Rattache a"), sinon DG.
        for d in rows["Departements"]:
            dep = dept_by_name.get(_norm_name(self._get(d, "nom")))
            if not dep:
                continue
            parent = dept_by_name.get(_norm_name(self._get(d, "rattache")))
            dep.parent = parent or dg
            dep.save(update_fields=["parent"])
        self.stdout.write(f"  Departements : {len(dept_by_name)} (dont Direction Generale)")

        # ---- 2. Employes + comptes ----
        taken = set(User.objects.values_list("username", flat=True))
        self.user_by_name = {}
        self.user_by_last = {}
        emp_specs = []  # (employee, manager_name) pour 2e passe
        n_emp = skipped = 0
        for d in rows["Employes"]:
            first = self._gs(d, "prenom")
            last = self._gs(d, "nom")
            email = self._gs(d, "email")
            if not (first or last):
                continue
            if email.lower() in EXAMPLE_EMAILS:   # ligne d'exemple laissee dans le modele
                skipped += 1
                continue
            role = ROLE_MAP.get(_ascii(self._get(d, "role")), Role.EMPLOYE)
            username = self._username_for(first, last, taken)
            user, created = User.objects.get_or_create(
                username=username,
                defaults=dict(first_name=first, last_name=last, email=email,
                              role=role, phone=self._gs(d, "telephone"),
                              must_change_password=True))
            if created:
                user.set_password(self.pwd)
                user.save()
            else:
                user.first_name, user.last_name, user.role = first, last, role
                user.email = email or user.email
                user.save()
            self.user_by_name[_norm_name(f"{first} {last}")] = user
            self.user_by_last.setdefault(_ascii(last), []).append(user)

            emp = Employee.objects.filter(user=user).first() or Employee.objects.create(user=user)
            emp.gender = GENDER_MAP.get(_ascii(self._get(d, "sexe")), "")
            emp.contract_type = CONTRACT_MAP.get(_ascii(self._get(d, "contrat", "type de contrat")),
                                                 Employee.Contract.CDI)
            hd = _to_date(self._get(d, "embauche"))
            if hd:
                emp.hire_date = hd
            jour = _to_int(self._get(d, "jour"), 0)
            mois = _to_int(self._get(d, "mois"), 0)
            if jour and mois:
                try:
                    emp.birth_date = date(2000, mois, jour)
                except ValueError:
                    pass
            emp.emergency_contact = self._gs(d, "urgence - nom", "contact urgence - nom")
            emp.emergency_contact_phone = self._gs(d, "urgence - tel", "contact urgence - tel")
            emp.status = Employee.Status.ACTIVE
            emp.save()

            # Departements (multi) + fonction/poste
            depts = [dept_by_name[_norm_name(n)] for n in _split(self._get(d, "departement"))
                     if _norm_name(n) in dept_by_name]
            if depts:
                emp.departments.set(depts)
            poste = self._gs(d, "fonction", "poste")
            if poste:
                dep0 = depts[0] if depts else None
                pos = (Position.objects.filter(title__iexact=poste).first()
                       or Position.objects.create(title=poste, department=dep0))
                emp.positions.add(pos)
            emp_specs.append((emp, self._get(d, "hierarchique", "responsable hierarchique")))
            n_emp += 1
        msg = f"  Employes/comptes : {n_emp}"
        if skipped:
            msg += f" ({skipped} ligne(s) d'exemple ignoree(s))"
        self.stdout.write(msg)

        # ---- 2b. Responsables hierarchiques (2e passe) ----
        for emp, mgr_name in emp_specs:
            u = self._resolve_person(mgr_name) if mgr_name else None
            mgr_emp = Employee.objects.filter(user=u).first() if u else None
            if mgr_emp and mgr_emp.pk != emp.pk:
                emp.manager = mgr_emp
                emp.save(update_fields=["manager"])

        # ---- 2c. Responsables de departement + Direction Generale (CEO) ----
        for d in rows["Departements"]:
            dep = dept_by_name.get(_norm_name(self._get(d, "nom")))
            resp = self._resolve_person(self._get(d, "responsable"))
            if dep and resp:
                dep.manager = resp
                dep.save(update_fields=["manager"])
        ceo = User.objects.filter(role=Role.CEO).first()
        if ceo:
            dg.manager = ceo
            dg.save(update_fields=["manager"])
            ceo_emp = Employee.objects.filter(user=ceo).first()
            if ceo_emp and not ceo_emp.departments.exists():
                ceo_emp.departments.set([dg])

        # ---- 3. Clients ----
        from business.models import Client as BClient
        self.client_by_name = {}
        for d in rows["Clients"]:
            name = self._gs(d, "nom")
            if not name:
                continue
            kind = CLIENT_KIND_MAP.get(_ascii(self._get(d, "type")), BClient.Kind.CLIENT)
            obj, _ = BClient.objects.get_or_create(name=name, defaults=dict(kind=kind))
            obj.kind = kind
            obj.contact_name = self._gs(d, "contact")
            obj.email = self._gs(d, "email")
            obj.phone = self._gs(d, "telephone")
            obj.sector = self._gs(d, "secteur")
            obj.save()
            self.client_by_name[_norm_name(name)] = obj
        self.stdout.write(f"  Clients : {len(self.client_by_name)}")

        # ---- 4. Projets ----
        from projects.models import Project
        n_proj = sk_proj = 0
        for d in rows["Projets"]:
            name = self._gs(d, "nom")
            if not name:
                continue
            if _norm_name(name) in EXAMPLE_PROJECTS:   # projet d'exemple laisse dans le modele
                sk_proj += 1
                continue
            obj, _ = Project.objects.get_or_create(name=name, defaults={})
            obj.kind = PROJ_KIND_MAP.get(_ascii(self._get(d, "type")), Project.Kind.OTHER)
            obj.status = PROJ_STATUS_MAP.get(_ascii(self._get(d, "statut")), Project.Status.PLANNED)
            obj.department = dept_by_name.get(_norm_name(self._get(d, "departement")))
            obj.client = self.client_by_name.get(_norm_name(self._get(d, "client")))
            obj.manager = self._resolve_person(self._get(d, "chef"))
            obj.start_date = _to_date(self._get(d, "debut"))
            obj.deadline = _to_date(self._get(d, "echeance"))
            obj.budget = _to_int(self._get(d, "budget (fcfa)", "budget"))
            obj.spent = _to_int(self._get(d, "consomme"))
            obj.location = self._gs(d, "lieu")
            obj.event_date = _to_date(self._get(d, "date evenement", "evenement (aaaa"))
            obj.save()
            team = [u for u in (self._resolve_person(n) for n in _split(self._get(d, "equipe"))) if u]
            if team:
                obj.team.set(team)
            n_proj += 1
        msg = f"  Projets : {n_proj}"
        if sk_proj:
            msg += f" ({sk_proj} exemple ignore)"
        self.stdout.write(msg)

        # ---- 5. Magasin ----
        if self.magasin_simple:
            self._import_magasin_simple()
        else:
            self._import_magasin_modele(rows["Magasin"])

    def _import_magasin_simple(self):
        """Magasin depuis le fichier separe : MATERIELS | QUANTITE | ACTION | PHOTO."""
        from stock.models import StockItem
        n = 0
        for d in self.magasin_simple:
            name = (self._get(d, "materiel", "materiels", "designation") or "")
            name = str(name).strip()
            if not name:
                continue
            action = (self._get(d, "action") or "")
            action = str(action).strip()
            photo = self._get(d, "photo")
            photo = str(photo).strip() if photo not in (None, "") else ""
            a = _ascii(action)
            if "repar" in a or "verifier" in a or "recharger" in a:
                etat = "REPAIR"
            elif "hors service" in a or "hs" == a:
                etat = "OUT_OF_SERVICE"
            else:
                etat = "USED"
            desc = action
            if photo:
                desc = (desc + f" — {photo}").strip(" —")
            obj = StockItem(
                name=name,
                category="OTHER",
                quantity=_to_int(self._get(d, "quantite", "quantité", "qte")),
                status=etat,
                description=desc,
            )
            obj.save()  # mat_id auto (MAT-xxx)
            n += 1
        self.stdout.write(f"  Articles magasin (fichier separe) : {n}")

    def _import_magasin_modele(self, rows):
        """Magasin depuis l'onglet 'Magasin' du modele (colonnes completes)."""
        from stock.models import StockItem, StockSupplier
        n = 0
        for d in rows:
            name = self._gs(d, "designation")
            if not name:
                continue
            sup = None
            sup_name = self._gs(d, "fournisseur")
            if sup_name:
                sup, _ = StockSupplier.objects.get_or_create(name=sup_name)
            mat = self._gs(d, "id materiel", "id mat") or None
            obj = (StockItem.objects.filter(mat_id=mat).first() if mat else None) \
                or StockItem(mat_id=mat)
            obj.mat_id = mat
            obj.name = name
            obj.brand_model = self._gs(d, "marque")
            obj.serial_number = self._gs(d, "serie")
            obj.category = STOCK_CAT_MAP.get(_ascii(self._get(d, "categorie")), "OTHER")
            obj.quantity = _to_int(self._get(d, "quantite"))
            obj.min_quantity = _to_int(self._get(d, "seuil"))
            obj.status = STOCK_STATE_MAP.get(_ascii(self._get(d, "etat")), "GOOD")
            obj.location = self._gs(d, "localisation")
            obj.estimated_value = _to_int(self._get(d, "valeur")) or None
            obj.supplier = sup
            obj.save()
            n += 1
        self.stdout.write(f"  Articles magasin : {n}")

    # ------------------------------------------------------------------ #
    def _dry_run(self, rows):
        """Verifie le mapping sans rien ecrire et signale les valeurs non reconnues."""
        self.stdout.write(self.style.MIGRATE_HEADING("\nSimulation (--dry-run) — aucune ecriture :"))
        warn = 0
        for d in rows["Employes"]:
            role_raw = self._get(d, "role")
            if role_raw and _ascii(role_raw) not in ROLE_MAP:
                warn += 1
                self.stdout.write(self.style.WARNING(
                    f"  [Employes] role non reconnu: '{role_raw}' -> defaut Employe"))
        for d in rows["Projets"]:
            st = self._get(d, "statut")
            if st and _ascii(st) not in PROJ_STATUS_MAP:
                warn += 1
                self.stdout.write(self.style.WARNING(
                    f"  [Projets] statut non reconnu: '{st}' -> defaut A demarrer"))
        for d in rows["Magasin"]:
            cat = self._get(d, "categorie")
            if cat and _ascii(cat) not in STOCK_CAT_MAP:
                warn += 1
                self.stdout.write(self.style.WARNING(
                    f"  [Magasin] categorie non reconnue: '{cat}' -> defaut Autre"))
        self.stdout.write(self.style.SUCCESS(
            f"\nSimulation OK — {warn} avertissement(s). "
            "Relance sans --dry-run (ajoute --flush pour remplacer la demo)."))
